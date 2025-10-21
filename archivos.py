import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
import openpyxl
from modelos import registro

# Crear carpeta reportes si no existe
if not os.path.exists("reportes"):
    os.makedirs("reportes")

def guardar_json(registro_obj: registro, archivo="datos.json"):
    with open(archivo, "w", encoding="utf-8") as f:
        json.dump(registro_obj.exportar_clase(), f, indent=4, ensure_ascii=False)
    print("Datos guardados correctamente ✔")

def cargar_json(archivo="datos.json"):
    """
    Carga datos desde archivo JSON.
    
    Args:
        archivo (str): Nombre del archivo JSON
        
    Returns:
        registro: Objeto registro reconstruido
    """
    try:
        with open(archivo, "r", encoding="utf-8") as f:
            data = json.load(f)
        return registro.importar_clase(data)
    except FileNotFoundError:
        print(f"Archivo {archivo} no encontrado")
        return None
    except json.JSONDecodeError:
        print(f"Error: el archivo {archivo} no es un JSON válido")
        return None
    except Exception as e:
        print(f"Error al cargar datos: {e}")
        return None
def exportar_a_excel(dataframe, nombre_archivo):
    """
    Exporta un DataFrame a archivo Excel.
    
    Args:
        dataframe (pd.DataFrame): DataFrame a exportar
        nombre_archivo (str): Nombre del archivo (sin extensión)
    """
    
    # Crear carpeta reportes si no existe
    if not os.path.exists("reportes"):
        os.makedirs("reportes")
    
    if not dataframe.empty:
        ruta = f"reportes/{nombre_archivo}.xlsx"
        dataframe.to_excel(ruta, index=False)
        print(f"Reporte exportado a: {ruta}")
    else:
        print("No hay datos para exportar")
def exportar_a_excel(dataframe, nombre_archivo):
    """
    Exporta un DataFrame a archivo Excel con formato mejorado.
    
    Args:
        dataframe (pd.DataFrame): DataFrame a exportar
        nombre_archivo (str): Nombre del archivo (sin extensión)
    """
    import os
    from openpyxl.styles import Border, Side, Alignment, Font
    
    # Crear carpeta reportes si no existe
    if not os.path.exists("reportes"):
        os.makedirs("reportes")
    
    if dataframe.empty:
        print("No hay datos para exportar")
        return
    
    ruta = f"reportes/{nombre_archivo}.xlsx"
    
    # Exportar a Excel
    dataframe.to_excel(ruta, index=False, sheet_name="Reporte")
    
    # Aplicar formato
    from openpyxl import load_workbook
    wb = load_workbook(ruta)
    ws = wb.active
    
    # Bordes y alineación
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formatear encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Formatear datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    wb.save(ruta)
    print(f"Reporte exportado a: {ruta}")
def cargar_desde_excel(archivo="datos.xlsx"):
    """
    Carga pacientes y evoluciones desde un Excel.
    
    Args:
        archivo (str): Nombre del archivo Excel (por defecto datos.xlsx)
        
    Returns:
        registro: Objeto registro reconstruido
    """
    import pandas as pd
    from modelos import registro, paciente, evolucion
    from datetime import datetime
    
    try:
        # Leer hojas
        df_pacientes = pd.read_excel(archivo, sheet_name="Pacientes")
        df_evoluciones = pd.read_excel(archivo, sheet_name="Evoluciones")
        
        # Crear registro
        reg = registro()
        
        # Cargar pacientes
        for _, row in df_pacientes.iterrows():
            cedula = int(row["Cédula"])
            nombre = row["Nombre"]
            apellido = row["Apellido"]
            
            p = paciente(cedula, nombre, apellido)
            reg.agregar_paciente(p)
        
        # Cargar evoluciones
        for _, row in df_evoluciones.iterrows():
            cedula = int(row["Cédula"])
            fecha = pd.to_datetime(row["Fecha"]).date()
            hora = pd.to_datetime(row["Hora"]).time()
            contenido = row["Contenido"]
            
            p = reg.obtener_paciente(cedula)
            if p is not None:
                ev = evolucion(fecha, hora, contenido)
                ev.retraso = {
                    "dias": int(row["Retraso (días)"]),
                    "horas": int(row["Retraso (horas)"]),
                    "minutos": int(row["Retraso (minutos)"])
                }
                p.agregar_evolucion(ev)
        
        print("Información cargada correctamente desde Excel")
        return reg
        
    except FileNotFoundError:
        print(f"Archivo {archivo} no encontrado")
        return None
    except Exception as e:
        print(f"Error al cargar Excel: {e}")
        return None
def exportar_sistema_completo_excel(registro_obj: registro, archivo="datos.xlsx"):
    """
    Exporta toda la información del sistema a un Excel.
    Incluye todas las hojas necesarias para cargar después.
    
    Args:
        registro_obj: Objeto registro del sistema
        archivo: Nombre del archivo (por defecto datos.xlsx)
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font
    import openpyxl
    
    # Preparar datos de pacientes
    datos_pacientes = []
    for paciente in registro_obj.pacientes.values():
        datos_pacientes.append({
            "Cédula": paciente.cedula,
            "Nombre": paciente.nombre,
            "Apellido": paciente.apellido
        })
    
    # Preparar datos de evoluciones
    datos_evoluciones = []
    for paciente in registro_obj.pacientes.values():
        for ev in paciente.evoluciones:
            datos_evoluciones.append({
                "Cédula": paciente.cedula,
                "Fecha": ev.fecha,
                "Hora": ev.hora,
                "Contenido": ev.contenido,
                "Retraso (días)": ev.retraso["dias"],
                "Retraso (horas)": ev.retraso["horas"],
                "Retraso (minutos)": ev.retraso["minutos"]
            })
    
    # Crear DataFrames
    df_pacientes = pd.DataFrame(datos_pacientes)
    df_evoluciones = pd.DataFrame(datos_evoluciones)
    
    # Exportar a Excel con múltiples hojas
    with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
        df_pacientes.to_excel(writer, sheet_name='Pacientes', index=False)
        df_evoluciones.to_excel(writer, sheet_name='Evoluciones', index=False)
    
    # Aplicar formato
    wb = load_workbook(archivo)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Formatear encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatear datos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2
    
    wb.save(archivo)
    print(f"Sistema completo exportado a: {archivo}")